# _*_ coding: utf-8 _*_
# @Time : 2021/11/17 10:14
# @Auth: GLK
# @File: config.py
# @Software :PyCharm
import os, time
import xlrd, json
from util import BASE_DIR
import logging
import pandas as pd
from pandas import DataFrame
from openpyxl import load_workbook


class Get_Test_Data():
    def __init__(self):
        self.lists = []
        self.dicts = {}

    def read_excel(self, filepath):
        self.x1 = xlrd.open_workbook(filepath)
        self.sheet = self.x1.sheet_by_name("Sheet1")
        self.keys = self.sheet.row_values(0)
        for i in range(1, self.sheet.nrows):
            row_content = []
            for j in range(self.sheet.ncols):
                ctype = self.sheet.cell(i, j).ctype  # 判断表格数据类型
                cell = self.sheet.cell_value(i, j)
                if ctype == 2 and cell % 1 == 0:  # 判断为int  去除表格为空的
                    cell = int(cell)
                    # row_content.append(cell)
                elif ctype == 0:  # 新增判断exce单元格为空的情况
                    continue
                row_content.append(cell)

            self.dicts = dict(zip(self.keys, row_content))
            self.lists.append(self.dicts)
        return self.lists


def write_excel(path, actual_code, actual_status, result):
    data = pd.read_excel(path)
    data["actual_code"] = actual_code
    data["actual_status"] = actual_status
    data["result"] = result
    DataFrame(data).to_excel(path, sheet_name="Sheet1", index=False, header=True)


def get_data(filepath):
    result_list = []
    g = Get_Test_Data()
    for data_list in g.read_excel(filepath):
        per_data = list(data_list.values())
        for i in per_data:
            if type(i) == str and i.startswith("{"):
                idx = per_data.index(i)
                per_data[idx] = eval(i)
        result_list.append(per_data)
    return result_list


class Logger():
    def __init__(self, logger):
        ''' 指定保存日志的文件路径，日志级别，以及调用文件
            将日志存入到指定的文件中
        :param logger:
        '''
        # 创建一个Logger
        self.logger = logging.getLogger(logger)
        self.logger.setLevel(logging.DEBUG)
        # 创建一个handler,用于写入日志文件
        rq = time.strftime('%Y%m%d%H%M', time.localtime(time.time()))
        log_path = os.path.dirname(os.getcwd()) + '/test-framework1/Logs/'  # 是获取你文件的路径
        log_path = BASE_DIR + "/log/"  # 是获取日志文件的路径
        log_name = log_path + rq + '.log'
        fh = logging.FileHandler(log_name)
        fh.setLevel(logging.INFO)
        # 创建一个handler，用于输出到控制台
        ch = logging.StreamHandler()
        ch.setLevel(logging.INFO)
        # 定义Handler的输出格式
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        fh.setFormatter(formatter)
        ch.setFormatter(formatter)
        # 给logger添加handler
        self.logger.addHandler(fh)
        self.logger.addHandler(ch)

    def getlog(self):
        return self.logger


class Clear_excel():
    """
    清除excel最后三行的数据
    """
    def __init__(self):
        self.nrows = 0
        self.ncols = 0

    def get_nrows(self, path):
        sheet = xlrd.open_workbook(path)
        sheet1 = sheet.sheet_by_name("Sheet1")
        self.nrows = sheet1.nrows
        self.ncols = sheet1.ncols

        wb = load_workbook(path)
        sheets = wb["Sheet1"]
        for i in range(2, self.nrows + 1):
            sheets.cell(row=i, column=self.ncols, value="")
        for i in range(2, self.nrows + 1):
            sheets.cell(row=i, column=self.ncols - 1, value="")
        for i in range(2, self.nrows + 1):
            sheets.cell(row=i, column=self.ncols - 2, value="")
        print(self.ncols, self.nrows)
        wb.save(path)

if __name__ == "__main__":
    path = BASE_DIR
    filepath = path + "\\" + "database" + "\\" + "login_data.xlsx"
    list=get_data(filepath)
    print(list)
    print(list[:-3])
# for i in get_data(filepath):
#     print(i)
